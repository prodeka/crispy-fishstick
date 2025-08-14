import openpyxl
import argparse
import os
from collections import defaultdict

"""
Script d'extraction amélioré des formules Excel vers Markdown avec options :
- Filtre les cellules vides (None, "", espaces)
- Affiche différemment les formules et les valeurs
- Options de filtrage avancées
- Génère des rapports plus propres

Usage :
    python extract_excel_formulas_improved.py --excel Reseaux_2.xlsx --output output_dir [--formulas-only] [--values-only] [--non-empty-only]

Options :
    --excel          Chemin du fichier Excel à analyser
    --output         Dossier de sortie pour les fichiers Markdown
    --formulas-only  Extraire uniquement les formules
    --values-only    Extraire uniquement les valeurs (pas les formules)
    --non-empty-only Extraire uniquement les cellules non vides
    --all            Extraire toutes les formules (pas seulement uniques)
    --preview N      Nombre de formules à afficher dans l'aperçu rapide (défaut : 10)
    --unique         Extraire uniquement les formules uniques par colonne (défaut : activé)
    --split          Générer un fichier Markdown par feuille (défaut : activé)
    --sheet S        Extraire uniquement la feuille nommée S (optionnel)
    --max-rows N     Limiter le nombre de lignes analysées (optionnel)
    --min-value      Valeur minimale pour considérer une cellule comme non vide (défaut : 0)
"""

def is_empty_value(value, min_value=0):
    """Vérifie si une valeur est considérée comme vide"""
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    if isinstance(value, (int, float)):
        return value < min_value
    return False

def extract_formulas_improved(
    excel_path,
    output_dir,
    unique=True,
    preview_count=10,
    split=True,
    sheet_filter=None,
    max_rows=None,
    formulas_only=False,
    values_only=False,
    non_empty_only=True,
    min_value=0
):
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    os.makedirs(output_dir, exist_ok=True)
    preview_lines = []
    
    for sheet in wb.sheetnames:
        if sheet_filter and sheet != sheet_filter:
            continue
            
        ws = wb[sheet]
        md_lines = [f"# Extraction des formules Excel\n", f"## Feuille : {sheet}\n"]
        md_lines.append("| Ligne | Colonne | Cellule | Formule | Valeur calculée |")
        md_lines.append("|-------|---------|---------|---------|-----------------|")
        
        valeur_lines = [f"# Valeurs des cellules Excel\n", f"## Feuille : {sheet}\n"]
        valeur_lines.append("| Ligne | Colonne | Cellule | Type | Contenu | Valeur calculée |")
        valeur_lines.append("|-------|---------|---------|------|---------|-----------------|")
        
        found_formulas = False
        found_values = False
        headers = {}
        
        # Récupération des en-têtes
        for col in ws.iter_cols(1, ws.max_column, 1, 1):
            cell = col[0]
            if cell.value:
                headers[cell.column] = str(cell.value)
        
        unique_formulas = defaultdict(set)  # col_name -> set(formules)
        row_count = 0
        
        for row in ws.iter_rows(min_row=2):
            if max_rows and row_count >= max_rows:
                break
            row_count += 1
            row_name = row[0].value if row and row[0].value else f"Ligne {row[0].row}"
            
            for cell in row:
                col_name = headers.get(cell.column, f"Col {cell.column}")
                cell_address = cell.coordinate
                
                # Récupération de la valeur calculée
                try:
                    calculated_value = ws.parent[sheet][cell_address].value
                except:
                    calculated_value = None
                
                # Vérification si c'est une formule
                is_formula = (cell.data_type == 'f' or 
                            (cell.value and isinstance(cell.value, str) and cell.value.startswith('=')))
                
                # Vérification si la cellule n'est pas vide
                is_not_empty = not is_empty_value(calculated_value, min_value)
                
                # Extraction des formules
                if not values_only and is_formula:
                    found_formulas = True
                    formula = cell.value
                    
                    if unique:
                        if formula in unique_formulas[col_name]:
                            continue
                        unique_formulas[col_name].add(formula)
                    
                    line = f"| {row_name} | {col_name} | {cell_address} | `{formula}` | {calculated_value} |"
                    md_lines.append(line)
                    
                    if len(preview_lines) < preview_count:
                        preview_lines.append(line)
                
                # Extraction des valeurs
                if not formulas_only and is_not_empty:
                    found_values = True
                    
                    if is_formula:
                        cell_type = "**Formule**"
                        contenu = f"`{cell.value}`"
                    else:
                        cell_type = "Valeur"
                        contenu = str(calculated_value)
                    
                    valeur_line = f"| {row_name} | {col_name} | {cell_address} | {cell_type} | {contenu} | {calculated_value} |"
                    valeur_lines.append(valeur_line)
        
        # Messages si rien trouvé
        if not found_formulas and not values_only:
            md_lines.append("_Aucune formule trouvée sur cette feuille._")
        
        if not found_values and not formulas_only:
            valeur_lines.append("_Aucune valeur non vide trouvée sur cette feuille._")
        
        # Écriture des fichiers
        if split:
            if not formulas_only:
                out_path = os.path.join(output_dir, f"formules_{sheet}.md")
                with open(out_path, "w", encoding="utf-8") as f:
                    f.write('\n'.join(md_lines))
            
            if not values_only:
                out_val_path = os.path.join(output_dir, f"valeurs_{sheet}.md")
                with open(out_val_path, "w", encoding="utf-8") as f:
                    f.write('\n'.join(valeur_lines))
    
    # Fichier global (toutes feuilles)
    if not split:
        if not formulas_only:
            out_path = os.path.join(output_dir, "formules_extraites.md")
            with open(out_path, "w", encoding="utf-8") as f:
                f.write('\n'.join(md_lines))
    
    # Aperçu rapide
    if not values_only and preview_lines:
        preview_path = os.path.join(output_dir, "formules_apercu.md")
        with open(preview_path, "w", encoding="utf-8") as f:
            f.write("# Aperçu rapide des formules extraites\n\n")
            f.write("\n".join(preview_lines))

def main():
    parser = argparse.ArgumentParser(description="Extraction améliorée des formules Excel vers Markdown.")
    parser.add_argument('--excel', type=str, required=True, help='Fichier Excel à analyser')
    parser.add_argument('--output', type=str, required=True, help='Dossier de sortie pour les fichiers Markdown')
    parser.add_argument('--formulas-only', action='store_true', help='Extraire uniquement les formules')
    parser.add_argument('--values-only', action='store_true', help='Extraire uniquement les valeurs (pas les formules)')
    parser.add_argument('--non-empty-only', action='store_true', default=True, help='Extraire uniquement les cellules non vides (défaut)')
    parser.add_argument('--all', action='store_true', help='Extraire toutes les formules (pas seulement uniques)')
    parser.add_argument('--preview', type=int, default=10, help='Nombre de formules à afficher dans l\'aperçu rapide')
    parser.add_argument('--unique', action='store_true', help='Extraire uniquement les formules uniques (défaut)')
    parser.add_argument('--split', action='store_true', help='Générer un fichier Markdown par feuille (défaut)')
    parser.add_argument('--sheet', type=str, help='Nom de la feuille à extraire (optionnel)')
    parser.add_argument('--max-rows', type=int, help='Limiter le nombre de lignes analysées (optionnel)')
    parser.add_argument('--min-value', type=float, default=0, help='Valeur minimale pour considérer une cellule comme non vide')
    
    args = parser.parse_args()
    
    extract_formulas_improved(
        excel_path=args.excel,
        output_dir=args.output,
        unique=not args.all,
        preview_count=args.preview,
        split=args.split or not args.sheet,
        sheet_filter=args.sheet,
        max_rows=args.max_rows,
        formulas_only=args.formulas_only,
        values_only=args.values_only,
        non_empty_only=args.non_empty_only,
        min_value=args.min_value
    )

if __name__ == "__main__":
    main() 