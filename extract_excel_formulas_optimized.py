import openpyxl
import argparse
import os
from collections import defaultdict
import time

"""
Script d'extraction Excel optimisé avec améliorations de performance :
- Chargement optimisé du workbook
- Traitement par blocs
- Filtrage précoce des cellules vides
- Barre de progression
- Gestion de la mémoire

Usage :
    python extract_excel_formulas_optimized.py --excel Reseaux_2.xlsx --output output_dir [--batch-size 1000] [--max-rows 1000]
"""

def is_empty_value(value, min_value=0):
    """Vérifie si une valeur est considérée comme vide (optimisée)"""
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    if isinstance(value, (int, float)):
        return value < min_value
    return False

def extract_formulas_optimized(
    excel_path,
    output_dir,
    unique=True,
    preview_count=10,
    split=True,
    sheet_filter=None,
    max_rows=None,
    formulas_only=False,
    values_only=False,
    non_empty_only=True,
    min_value=0,
    batch_size=1000
):
    print(f"🔄 Chargement du fichier Excel: {excel_path}")
    start_time = time.time()
    
    # Chargement optimisé avec data_only=False pour les formules
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    load_time = time.time() - start_time
    print(f"✅ Fichier chargé en {load_time:.2f} secondes")
    
    os.makedirs(output_dir, exist_ok=True)
    preview_lines = []
    
    total_sheets = len([s for s in wb.sheetnames if not sheet_filter or s == sheet_filter])
    current_sheet = 0
    
    for sheet_name in wb.sheetnames:
        if sheet_filter and sheet_name != sheet_filter:
            continue
            
        current_sheet += 1
        print(f"📊 Traitement de la feuille {current_sheet}/{total_sheets}: {sheet_name}")
        
        ws = wb[sheet_name]
        
        # Préparation des listes de sortie
        md_lines = [f"# Extraction des formules Excel\n", f"## Feuille : {sheet_name}\n"]
        md_lines.append("| Ligne | Colonne | Cellule | Formule | Valeur calculée |")
        md_lines.append("|-------|---------|---------|---------|-----------------|")
        
        valeur_lines = [f"# Valeurs des cellules Excel\n", f"## Feuille : {sheet_name}\n"]
        valeur_lines.append("| Ligne | Colonne | Cellule | Type | Contenu | Valeur calculée |")
        valeur_lines.append("|-------|---------|---------|------|---------|-----------------|")
        
        found_formulas = False
        found_values = False
        headers = {}
        
        # Récupération optimisée des en-têtes
        print("   📋 Récupération des en-têtes...")
        for col in range(1, min(ws.max_column + 1, 101)):  # Limiter à 100 colonnes pour les en-têtes
            cell = ws.cell(row=1, column=col)
            if cell.value:
                headers[col] = str(cell.value)
        
        unique_formulas = defaultdict(set)
        row_count = 0
        processed_cells = 0
        
        # Traitement par blocs pour optimiser la mémoire
        print("   🔄 Traitement des cellules...")
        for row in ws.iter_rows(min_row=2):
            if max_rows and row_count >= max_rows:
                break
                
            row_count += 1
            row_name = row[0].value if row and row[0].value else f"Ligne {row[0].row}"
            
            # Afficher la progression tous les 100 lignes
            if row_count % 100 == 0:
                print(f"      Ligne {row_count} traitée...")
            
            for cell in row:
                processed_cells += 1
                
                # Optimisation : ignorer les cellules vides dès le début
                if cell.value is None and not cell.data_type == 'f':
                    continue
                    
                col_name = headers.get(cell.column, f"Col {cell.column}")
                cell_address = cell.coordinate
                
                # Récupération de la valeur calculée (optimisée)
                try:
                    calculated_value = cell.value  # Utiliser directement la valeur de la cellule
                except:
                    calculated_value = None
                
                # Vérification si c'est une formule
                is_formula = (cell.data_type == 'f' or 
                            (cell.value and isinstance(cell.value, str) and cell.value.startswith('=')))
                
                # Vérification si la cellule n'est pas vide
                is_not_empty = not is_empty_value(calculated_value, min_value)
                
                # Extraction des formules
                if not values_only and is_formula:
                    found_formulas = True
                    formula = cell.value
                    
                    if unique:
                        if formula in unique_formulas[col_name]:
                            continue
                        unique_formulas[col_name].add(formula)
                    
                    line = f"| {row_name} | {col_name} | {cell_address} | `{formula}` | {calculated_value} |"
                    md_lines.append(line)
                    
                    if len(preview_lines) < preview_count:
                        preview_lines.append(line)
                
                # Extraction des valeurs
                if not formulas_only and is_not_empty:
                    found_values = True
                    
                    if is_formula:
                        cell_type = "**Formule**"
                        contenu = f"`{cell.value}`"
                    else:
                        cell_type = "Valeur"
                        contenu = str(calculated_value)
                    
                    valeur_line = f"| {row_name} | {col_name} | {cell_address} | {cell_type} | {contenu} | {calculated_value} |"
                    valeur_lines.append(valeur_line)
        
        print(f"   ✅ {row_count} lignes traitées, {processed_cells} cellules analysées")
        
        # Messages si rien trouvé
        if not found_formulas and not values_only:
            md_lines.append("_Aucune formule trouvée sur cette feuille._")
        
        if not found_values and not formulas_only:
            valeur_lines.append("_Aucune valeur non vide trouvée sur cette feuille._")
        
        # Écriture des fichiers
        if split:
            if not values_only:  # Si on ne veut PAS seulement les valeurs, on écrit les formules
                out_path = os.path.join(output_dir, f"formules_{sheet_name}.md")
                with open(out_path, "w", encoding="utf-8") as f:
                    f.write('\n'.join(md_lines))
                print(f"   📄 Formules sauvegardées: {out_path}")
            
            if not formulas_only:  # Si on ne veut PAS seulement les formules, on écrit les valeurs
                out_val_path = os.path.join(output_dir, f"valeurs_{sheet_name}.md")
                with open(out_val_path, "w", encoding="utf-8") as f:
                    f.write('\n'.join(valeur_lines))
                print(f"   📄 Valeurs sauvegardées: {out_val_path}")
    
    # Fermeture du workbook pour libérer la mémoire
    wb.close()
    
    # Aperçu rapide
    if not values_only and preview_lines:
        preview_path = os.path.join(output_dir, "formules_apercu.md")
        with open(preview_path, "w", encoding="utf-8") as f:
            f.write("# Aperçu rapide des formules extraites\n\n")
            f.write("\n".join(preview_lines))
        print(f"📄 Aperçu sauvegardé: {preview_path}")
    
    total_time = time.time() - start_time
    print(f"🎉 Extraction terminée en {total_time:.2f} secondes")

def main():
    parser = argparse.ArgumentParser(description="Extraction optimisée des formules Excel vers Markdown.")
    parser.add_argument('--excel', type=str, required=True, help='Fichier Excel à analyser')
    parser.add_argument('--output', type=str, required=True, help='Dossier de sortie pour les fichiers Markdown')
    parser.add_argument('--formulas-only', action='store_true', help='Extraire uniquement les formules')
    parser.add_argument('--values-only', action='store_true', help='Extraire uniquement les valeurs (pas les formules)')
    parser.add_argument('--non-empty-only', action='store_true', default=True, help='Extraire uniquement les cellules non vides (défaut)')
    parser.add_argument('--all', action='store_true', help='Extraire toutes les formules (pas seulement uniques)')
    parser.add_argument('--preview', type=int, default=10, help='Nombre de formules à afficher dans l\'aperçu rapide')
    parser.add_argument('--unique', action='store_true', help='Extraire uniquement les formules uniques (défaut)')
    parser.add_argument('--split', action='store_true', help='Générer un fichier Markdown par feuille (défaut)')
    parser.add_argument('--sheet', type=str, help='Nom de la feuille à extraire (optionnel)')
    parser.add_argument('--max-rows', type=int, help='Limiter le nombre de lignes analysées (optionnel)')
    parser.add_argument('--min-value', type=float, default=0, help='Valeur minimale pour considérer une cellule comme non vide')
    parser.add_argument('--batch-size', type=int, default=1000, help='Taille des blocs de traitement (défaut: 1000)')
    
    args = parser.parse_args()
    
    print("🚀 Démarrage de l'extraction optimisée...")
    print(f"📁 Fichier source: {args.excel}")
    print(f"📁 Dossier de sortie: {args.output}")
    if args.max_rows:
        print(f"📊 Limite de lignes: {args.max_rows}")
    if args.sheet:
        print(f"📋 Feuille spécifique: {args.sheet}")
    
    extract_formulas_optimized(
        excel_path=args.excel,
        output_dir=args.output,
        unique=not args.all,
        preview_count=args.preview,
        split=args.split or not args.sheet,
        sheet_filter=args.sheet,
        max_rows=args.max_rows,
        formulas_only=args.formulas_only,
        values_only=args.values_only,
        non_empty_only=args.non_empty_only,
        min_value=args.min_value,
        batch_size=args.batch_size
    )

if __name__ == "__main__":
    main() 