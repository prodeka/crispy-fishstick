import openpyxl
import argparse
import os
from collections import defaultdict

"""
Script d'extraction des formules Excel vers Markdown avec options :
- Génère un fichier Markdown par feuille (par défaut)
- Filtre les formules uniques par colonne (par défaut)
- Génère un aperçu rapide (par défaut : 10 premières formules)
- Ajoute le nom de la ligne si présent (première colonne)
- Options ajustables via la ligne de commande

Usage :
    python extract_excel_formulas.py --excel Reseaux_2.xlsx --output output_dir [--all] [--preview 10] [--unique] [--split]

Options :
    --excel      Chemin du fichier Excel à analyser
    --output     Dossier de sortie pour les fichiers Markdown
    --all        Extraire toutes les formules (pas seulement uniques)
    --preview N  Nombre de formules à afficher dans l'aperçu rapide (défaut : 10)
    --unique     Extraire uniquement les formules uniques par colonne (défaut : activé)
    --split      Générer un fichier Markdown par feuille (défaut : activé)
    --sheet S    Extraire uniquement la feuille nommée S (optionnel)
    --max-rows N Limiter le nombre de lignes analysées (optionnel)
"""

def extract_formulas(
    excel_path,
    output_dir,
    unique=True,
    preview_count=10,
    split=True,
    sheet_filter=None,
    max_rows=None
):
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    os.makedirs(output_dir, exist_ok=True)
    preview_lines = []
    for sheet in wb.sheetnames:
        if sheet_filter and sheet != sheet_filter:
            continue
        ws = wb[sheet]
        md_lines = [f"# Extraction des formules Excel\n", f"## Feuille : {sheet}\n"]
        md_lines.append("| Ligne | Colonne | Cellule | Formule | Valeur calculée |")
        md_lines.append("|-------|---------|---------|---------|-----------------|")
        valeur_lines = [f"# Valeurs des cellules Excel\n", f"## Feuille : {sheet}\n"]
        valeur_lines.append("| Ligne | Colonne | Cellule | Type | Contenu | Valeur calculée |")
        valeur_lines.append("|-------|---------|---------|------|---------|-----------------|")
        found = False
        headers = {}
        for col in ws.iter_cols(1, ws.max_column, 1, 1):
            cell = col[0]
            if cell.value:
                headers[cell.column] = str(cell.value)
        unique_formulas = defaultdict(set)  # col_name -> set(formules)
        row_count = 0
        for row in ws.iter_rows(min_row=2):
            if max_rows and row_count >= max_rows:
                break
            row_count += 1
            row_name = row[0].value if row and row[0].value else f"Ligne {row[0].row}"
            for cell in row:
                col_name = headers.get(cell.column, f"Col {cell.column}")
                cell_address = cell.coordinate
                value = ws.parent[sheet][cell_address].value
                # Génération du fichier formules
                if cell.data_type == 'f' or (cell.value and isinstance(cell.value, str) and cell.value.startswith('=')):
                    found = True
                    formula = cell.value
                    if unique:
                        if formula in unique_formulas[col_name]:
                            continue
                        unique_formulas[col_name].add(formula)
                    line = f"| {row_name} | {col_name} | {cell_address} | `{formula}` | {value} |"
                    md_lines.append(line)
                    if len(preview_lines) < preview_count:
                        preview_lines.append(line)
                # Génération du fichier valeurs (pour toutes les cellules non vides)
                if value is not None and value != "":
                    if cell.data_type == 'f' or (cell.value and isinstance(cell.value, str) and cell.value.startswith('=')):
                        cell_type = "Formule"
                        contenu = cell.value
                    else:
                        cell_type = "Valeur"
                        contenu = value
                    valeur_line = f"| {row_name} | {col_name} | {cell_address} | {cell_type} | {contenu} | {value} |"
                    valeur_lines.append(valeur_line)
        if not found:
            md_lines.append("_Aucune formule trouvée sur cette feuille._")
        if split:
            out_path = os.path.join(output_dir, f"formules_{sheet}.md")
            with open(out_path, "w", encoding="utf-8") as f:
                f.write('\n'.join(md_lines))
            out_val_path = os.path.join(output_dir, f"valeurs_{sheet}.md")
            with open(out_val_path, "w", encoding="utf-8") as f:
                f.write('\n'.join(valeur_lines))
    # Fichier global (toutes feuilles)
    if not split:
        out_path = os.path.join(output_dir, "formules_extraites.md")
        with open(out_path, "w", encoding="utf-8") as f:
            f.write('\n'.join(md_lines))
    # Aperçu rapide
    preview_path = os.path.join(output_dir, "formules_apercu.md")
    with open(preview_path, "w", encoding="utf-8") as f:
        f.write("# Aperçu rapide des formules extraites\n\n")
        f.write("\n".join(preview_lines))

def main():
    parser = argparse.ArgumentParser(description="Extraction avancée des formules Excel vers Markdown.")
    parser.add_argument('--excel', type=str, required=True, help='Fichier Excel à analyser')
    parser.add_argument('--output', type=str, required=True, help='Dossier de sortie pour les fichiers Markdown')
    parser.add_argument('--all', action='store_true', help='Extraire toutes les formules (pas seulement uniques)')
    parser.add_argument('--preview', type=int, default=10, help='Nombre de formules à afficher dans l\'aperçu rapide')
    parser.add_argument('--unique', action='store_true', help='Extraire uniquement les formules uniques (défaut)')
    parser.add_argument('--split', action='store_true', help='Générer un fichier Markdown par feuille (défaut)')
    parser.add_argument('--sheet', type=str, help='Nom de la feuille à extraire (optionnel)')
    parser.add_argument('--max-rows', type=int, help='Limiter le nombre de lignes analysées (optionnel)')
    args = parser.parse_args()
    extract_formulas(
        excel_path=args.excel,
        output_dir=args.output,
        unique=not args.all,
        preview_count=args.preview,
        split=args.split or not args.sheet,  # split par défaut sauf si une seule feuille
        sheet_filter=args.sheet,
        max_rows=args.max_rows
    )

if __name__ == "__main__":
    main()